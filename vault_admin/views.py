from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from vault.models import User, Category, Product, ProductVariant, VariantImage, Order, ReturnRequest, ItemReturnRequest, Wallet, ReferralReward, CategoryOffer, ReferralOffer, Coupon, CouponUsage, OrderItem
from django.contrib import messages
from django.core.files.base import ContentFile
from io import BytesIO
from django.db.models import Q, Count, Sum
from decimal import Decimal
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from django.template.loader import get_template
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa
from django.utils import timezone
from datetime import datetime, timedelta
from PIL import Image
import re
from django.utils import timezone
from zoneinfo import ZoneInfo 

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def user_management_page(request):
    query = request.GET.get('q', '').strip()
    status = request.GET.get('status', 'all')
    sort_order = request.GET.get('sort', 'desc')
    user_list = User.objects.filter(is_staff=False)

    if query:
        user_list = user_list.filter(
            Q(full_name__icontains=query) |
            Q(email__icontains=query) |
            Q(ph_number__icontains=query)
        )

    if status == 'active':
        user_list = user_list.filter(is_active=True)
    elif status == 'blocked':
        user_list = user_list.filter(is_active=False)

    if sort_order == 'asc':
        user_list = user_list.order_by('created_at')
    else:
        user_list = user_list.order_by('-created_at')

    paginator = Paginator(user_list, 10)
    page = request.GET.get('page')
    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    active_count = User.objects.filter(is_active=True, is_staff=False).count()
    blocked_count = User.objects.filter(is_active=False, is_staff=False).count()

    return render(request, 'user_management.html', {
        'users': users,
        'active_count': active_count,
        'blocked_count': blocked_count,
        'query': query,
        'status': status,
        'sort_order': sort_order,
    })

@login_required
def block_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if not user.is_staff:
        user.is_active = False
        user.save()
    return redirect('user_management')

@login_required
def unblock_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if not user.is_staff:
        user.is_active = True
        user.save()
    return redirect('user_management')

@login_required
@user_passes_test(lambda u: u.is_staff)
def category_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    categories = Category.objects.all()

    if status_filter == 'active':
        categories = categories.filter(is_deleted=False)
    elif status_filter == 'inactive':
        categories = categories.filter(is_deleted=True)

    if query:
        categories = categories.filter(Q(name__icontains=query) | Q(description__icontains=query))

    categories = categories.order_by('-created_at')
    total_categories = Category.objects.count()
    active_categories = Category.objects.filter(is_deleted=False).count()
    inactive_categories = Category.objects.filter(is_deleted=True).count()

    paginator = Paginator(categories, 10)
    page = request.GET.get('page')
    try:
        categories = paginator.page(page)
    except PageNotAnInteger:
        categories = paginator.page(1)
    except EmptyPage:
        categories = paginator.page(paginator.num_pages)

    return render(request, 'category_list.html', {
        'categories': categories,
        'query': query,
        'status_filter': status_filter,
        'total_categories': total_categories,
        'active_categories': active_categories,
        'inactive_categories': inactive_categories,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        errors = []

        if not name:
            errors.append("Enter category name")
        elif not re.match(r'^[a-zA-Z\s]+$', name):
            errors.append("Category name should only contain letters and spaces")
        elif len(name) < 2:
            errors.append("Category name should be at least 2 characters long")
        elif len(name) > 50:
            errors.append("Category name should not exceed 50 characters")

        if not image:
            errors.append("Category image is required")
        else:
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
            if image.content_type not in allowed_types:
                errors.append("Please upload a valid image file (JPG, PNG, GIF)")
            if image.size > 10 * 1024 * 1024:
                errors.append("Image size should be less than 10MB")

        if not errors and Category.objects.filter(name__iexact=name, is_deleted=False).exists():
            errors.append("A category with this name already exists")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'add_category.html', {
                'name': name,
                'description': description
            })

        try:
            Category.objects.create(name=name, description=description, image=image)
            messages.success(request, f"Category '{name}' added successfully")
            return redirect('category_list')
        except Exception as e:
            messages.error(request, "An error occurred while adding the category")
            return render(request, 'add_category.html', {
                'name': name,
                'description': description
            })

    return render(request, 'add_category.html')

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id, is_deleted=False)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        image = request.FILES.get('image')
        errors = []

        if not name:
            errors.append("Enter category name")
        elif not re.match(r'^[a-zA-Z\s]+$', name):
            errors.append("Category name should only contain letters and spaces")
        elif len(name) < 2:
            errors.append("Category name should be at least 2 characters long")
        elif len(name) > 50:
            errors.append("Category name should not exceed 50 characters")

        if not errors and Category.objects.filter(name__iexact=name, is_deleted=False).exclude(id=category_id).exists():
            errors.append("A category with this name already exists")

        if image:
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
            if image.content_type not in allowed_types:
                errors.append("Please upload a valid image file (JPG, PNG, GIF)")
            if image.size > 10 * 1024 * 1024:
                errors.append("Image size should be less than 10MB")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'edit_category.html', {
                'category': category,
                'name': name,
                'description': description
            })

        try:
            category.name = name
            category.description = description
            if image:
                category.image = image
            category.save()
            messages.success(request, f"Category '{name}' updated successfully")
            return redirect('category_list')
        except Exception as e:
            messages.error(request, "An error occurred while updating the category")
            return render(request, 'edit_category.html', {
                'category': category,
                'name': name,
                'description': description
            })

    return render(request, 'edit_category.html', {
        'category': category
    })
    
@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def category_offer_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    category_filter = request.GET.get('category', 'all')
    
    offers = CategoryOffer.objects.select_related('category').all()
    now_utc = timezone.now()  # timezone-aware UTC
    now = now_utc.astimezone(ZoneInfo("Asia/Kolkata"))
    
    if query:
        offers = offers.filter(
            Q(offer_name__icontains=query) |
            Q(category__name__icontains=query) |
            Q(description__icontains=query)
        )
    
    if status_filter == 'active':
        offers = offers.filter(is_active=True)
    elif status_filter == 'inactive':
        offers = offers.filter(is_active=False)
    elif status_filter == 'expired':
        offers = offers.filter(valid_until__lt=timezone.now())
    elif status_filter == 'upcoming':
        offers = offers.filter(valid_from__gt=timezone.now())
    
    if category_filter != 'all':
        try:
            category_filter = int(category_filter)
            offers = offers.filter(category_id=category_filter)
        except (ValueError, TypeError):
            category_filter = 'all'
    
    offers = offers.order_by('-created_at')
    
    categories = Category.objects.filter(is_deleted=False).order_by('name')
    
    paginator = Paginator(offers, 10)
    page = request.GET.get('page')
    
    try:
        offers = paginator.page(page)
    except PageNotAnInteger:
        offers = paginator.page(1)
    except EmptyPage:
        offers = paginator.page(paginator.num_pages)
    
    # Statistics
    total_offers = CategoryOffer.objects.count()
    active_offers = CategoryOffer.objects.filter(is_active=True).count()
    expired_offers = CategoryOffer.objects.filter(valid_until__lt=timezone.now()).count()
    
    context = {
        'offers': offers,
        'categories': categories,
        'query': query,
        'status_filter': status_filter,
        'category_filter': str(category_filter) if category_filter != 'all' else 'all',
        'total_offers': total_offers,
        'active_offers': active_offers,
        'expired_offers': expired_offers,
        'now':now
    }
    
    return render(request, 'offers/category_offer_list.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_category_offer(request):
    if request.method == 'POST':
        offer_name = request.POST.get('offer_name', '').strip()
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category')
        discount_percentage = request.POST.get('discount_percentage', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        
        errors = []
        now_utc = timezone.now()  # timezone-aware UTC
        now = now_utc.astimezone(ZoneInfo("Asia/Kolkata"))
        
        # Validate offer name
        if not offer_name:
            errors.append("Offer name is required")
        elif len(offer_name) < 3:
            errors.append("Offer name should be at least 3 characters long")
        elif len(offer_name) > 100:
            errors.append("Offer name should not exceed 100 characters")
        
        # Validate category
        if not category_id:
            errors.append("Please select a category")
        else:
            try:
                category = Category.objects.get(id=category_id, is_deleted=False)
            except Category.DoesNotExist:
                errors.append("Selected category does not exist")
        
        # Validate discount percentage
        if not discount_percentage:
            errors.append("Discount percentage is required")
        else:
            try:
                discount_percentage = float(discount_percentage)
                if discount_percentage <= 0 or discount_percentage > 100:
                    errors.append("Discount percentage must be between 0.01 and 100")
            except ValueError:
                errors.append("Please enter a valid discount percentage")
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = timezone.make_aware(datetime.strptime(valid_from, '%Y-%m-%dT%H:%M'))
                until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%dT%H:%M'))

                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                
                if until_date <= timezone.now():
                    errors.append("Valid until date must be in the future")
                    
            except ValueError:
                errors.append("Please enter valid dates")
       

        if errors:
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            for error in errors:
                messages.error(request, error)
            return render(request, 'offers/add_category_offer.html', {
                'categories': categories,
                'offer_name': offer_name,
                'description': description,
                'category_id': category_id,
                'discount_percentage': request.POST.get('discount_percentage', ''),
                'valid_from': valid_from,
                'valid_until': valid_until,
                'now': now,
            })
        
        try:
            CategoryOffer.objects.create(
                category=category,
                offer_name=offer_name,
                description=description,
                discount_percentage=discount_percentage,
                valid_from=from_date,
                valid_until=until_date,
                is_active=True
            )
            messages.success(request, f"Category offer '{offer_name}' created successfully")
            return redirect('category_offer_list')
        except Exception as e:
            messages.error(request, "An error occurred while creating the offer")
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            return render(request, 'offers/add_category_offer.html', {
                'categories': categories,
                'offer_name': offer_name,
                'description': description,
                'category_id': category_id,
                'discount_percentage': request.POST.get('discount_percentage', ''),
                'valid_from': valid_from,
                'valid_until': valid_until,
                'now': now,
            })
    
    categories = Category.objects.filter(is_deleted=False).order_by('name')
    return render(request, 'offers/add_category_offer.html', {'categories': categories})

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_category_offer(request, offer_id):
    offer = get_object_or_404(CategoryOffer, id=offer_id)
    
    if request.method == 'POST':
        offer_name = request.POST.get('offer_name', '').strip()
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category')
        discount_percentage = request.POST.get('discount_percentage', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        
        errors = []
        
        # Validate offer name
        if not offer_name:
            errors.append("Offer name is required")
        elif len(offer_name) < 3:
            errors.append("Offer name should be at least 3 characters long")
        elif len(offer_name) > 100:
            errors.append("Offer name should not exceed 100 characters")
        
        # Validate category
        if not category_id:
            errors.append("Please select a category")
        else:
            try:
                category = Category.objects.get(id=category_id, is_deleted=False)
            except Category.DoesNotExist:
                errors.append("Selected category does not exist")
        
        # Validate discount percentage
        if not discount_percentage:
            errors.append("Discount percentage is required")
        else:
            try:
                discount_percentage = float(discount_percentage)
                if discount_percentage <= 0 or discount_percentage > 100:
                    errors.append("Discount percentage must be between 0.01 and 100")
            except ValueError:
                errors.append("Please enter a valid discount percentage")
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = datetime.strptime(valid_from, '%Y-%m-%dT%H:%M')
                until_date = datetime.strptime(valid_until, '%Y-%m-%dT%H:%M')
                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                    
            except ValueError:
                errors.append("Please enter valid dates")
        
        if errors:
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            for error in errors:
                messages.error(request, error)
            return render(request, 'offers/edit_category_offer.html', {
                'offer': offer,
                'categories': categories,
            })
        
        try:
            offer.category = category
            offer.offer_name = offer_name
            offer.description = description
            offer.discount_percentage = discount_percentage
            offer.valid_from = from_date
            offer.valid_until = until_date
            offer.save()
            
            messages.success(request, f"Category offer '{offer_name}' updated successfully")
            return redirect('category_offer_list')
        except Exception as e:
            messages.error(request, "An error occurred while updating the offer")
    
    categories = Category.objects.filter(is_deleted=False).order_by('name')
    return render(request, 'offers/edit_category_offer.html', {
        'offer': offer,
        'categories': categories,
    })
    
@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_category_offer_status(request, offer_id):
    offer = get_object_or_404(CategoryOffer, id=offer_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'activate' and not offer.is_active:
            offer.is_active = True
            offer.save()
            messages.success(request, f"Category offer '{offer.offer_name}' has been activated")
        elif action == 'deactivate' and offer.is_active:
            offer.is_active = False
            offer.save()
            messages.success(request, f"Category offer '{offer.offer_name}' has been deactivated")
        else:
            messages.error(request, "Invalid action or offer status")
    
    return redirect('category_offer_list')

@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_category_status(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if request.method == 'POST':
        action = request.POST.get('action')
        category_name = category.name

        if action == 'deactivate' and not category.is_deleted:
            category.is_deleted = True
            category.save()
            messages.success(request, f"Category '{category_name}' has been deactivated successfully")
        elif action == 'activate' and category.is_deleted:
            category.is_deleted = False
            category.save()
            messages.success(request, f"Category '{category_name}' has been activated successfully")
        else:
            messages.error(request, "Invalid action or category status")
        return redirect('category_list')

    return render(request, 'toggle_category_status.html', {
        'category': category,
        
    })

def resize_and_crop_image(image_file, size=(800, 600)):
    try:
        img = Image.open(image_file)
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGB')

        img_ratio = img.width / img.height
        target_ratio = size[0] / size[1]

        if img_ratio > target_ratio:
            new_height = img.height
            new_width = int(new_height * target_ratio)
            left = (img.width - new_width) // 2
            img = img.crop((left, 0, left + new_width, new_height))
        else:
            new_width = img.width
            new_height = int(new_width / target_ratio)
            top = (img.height - new_height) // 2
            img = img.crop((0, top, new_width, top + new_height))

        img = img.resize(size, Image.Resampling.LANCZOS)
        output = BytesIO()
        img.save(output, format='JPEG', quality=85, optimize=True)
        output.seek(0)
        return ContentFile(output.read())
    except Exception as e:
        raise Exception(f"Error processing image: {str(e)}")

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def product_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    category_filter = request.GET.get('category', 'all')
    sort_order = request.GET.get('sort', 'desc')

    products = Product.objects.select_related('category').prefetch_related('variants__images').annotate(
        variant_count=Count('variants')
    )

    if status_filter == 'active':
        products = products.filter(is_deleted=False)
    elif status_filter == 'inactive':
        products = products.filter(is_deleted=True)

    if category_filter != 'all':
        try:
            category_filter = int(category_filter)
            products = products.filter(category_id=category_filter)
        except (ValueError, TypeError):
            pass

    if query:
        search_terms = re.findall(r'\w+', query.lower())
        if search_terms:
            search_query = Q()
            for term in search_terms:
                search_query |= (
                    Q(product_name__icontains=term) |
                    Q(product_description__icontains=term) |
                    Q(category__name__icontains=term)
                )
            products = products.filter(search_query)
        else:
            products = products.filter(
                Q(product_name__icontains=query) |
                Q(product_description__icontains=query) |
                Q(category__name__icontains=query)
            )

    if sort_order == 'asc':
        products = products.order_by('created_at')
    else:
        products = products.order_by('-created_at')

    categories = Category.objects.filter(is_deleted=False).order_by('name')
    paginator = Paginator(products, 10)
    page = request.GET.get('page')
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    total_products = Product.objects.count()
    active_products = Product.objects.filter(is_deleted=False).count()
    inactive_products = Product.objects.filter(is_deleted=True).count()

    return render(request, 'product_lists.html', {
        'products': products,
        'categories': categories,
        'query': query,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'sort_order': sort_order,
        'total_products': total_products,
        'active_products': active_products,
        'inactive_products': inactive_products,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_product(request):
    if request.method == 'POST':
        product_name = request.POST.get('product_name', '').strip()
        product_description = request.POST.get('product_description', '').strip()
        category_id = request.POST.get('category')
        price = request.POST.get('price', '').strip()
        product_offer = request.POST.get('product_offer', '0').strip()
        main_image = request.FILES.get('main_image')
        errors = []

        if not product_name:
            errors.append("Product name is required")
        elif len(product_name) < 2:
            errors.append("Product name should be at least 2 characters long")
        elif len(product_name) > 100:
            errors.append("Product name should not exceed 100 characters")
        elif Product.objects.filter(product_name__iexact=product_name, is_deleted=False).exists():
            errors.append("A product with this name already exists")

        if not category_id:
            errors.append("Please select a category")
        else:
            try:
                category = Category.objects.get(id=category_id, is_deleted=False)
            except Category.DoesNotExist:
                errors.append("Selected category does not exist")

        if not price:
            errors.append("Price is required")
        else:
            try:
                price = int(price)
                if price <= 0:
                    errors.append("Price must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid price")

        try:
            product_offer = float(product_offer)
            if product_offer < 0 or product_offer > 100:
                errors.append("Offer percentage must be between 0 and 100")
        except ValueError:
            errors.append("Please enter a valid offer percentage")

        if not main_image:
            errors.append("Main product image is required")
        else:
            if not main_image.content_type.startswith('image/'):
                errors.append("Please upload a valid image file for main image")
            if main_image.size > 10 * 1024 * 1024:
                errors.append("Main image should be less than 10MB")

        if errors:
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            for error in errors:
                messages.error(request, error)
            return render(request, 'add_product.html', {
                'categories': categories,
                'product_name': product_name,
                'product_description': product_description,
                'category_id': category_id,
                'price': request.POST.get('price', ''),
                'product_offer': request.POST.get('product_offer', ''),
            })

        try:
            processed_main_image = resize_and_crop_image(main_image)
            
            product = Product.objects.create(
                product_name=product_name,
                product_description=product_description,
                category=category,
                price=price,
                product_offer=product_offer
            )

            product.main_image.save(
                f"{product.product_name}_main.jpg",
                processed_main_image,
                save=True
            )

            messages.success(request, f"Product '{product_name}' added successfully. You can now add color variants.")
            return redirect('product_variants', product_id=product.id)
        except Exception as e:
            messages.error(request, "An error occurred while adding the product")
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            return render(request, 'add_product.html', {
                'categories': categories,
                'product_name': product_name,
                'product_description': product_description,
                'category_id': category_id,
                'price': request.POST.get('price', ''),
                'product_offer': request.POST.get('product_offer', ''),
            })

    categories = Category.objects.filter(is_deleted=False).order_by('name')
    return render(request, 'add_product.html', {'categories': categories})

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        product_name = request.POST.get('product_name', '').strip()
        product_description = request.POST.get('product_description', '').strip()
        category_id = request.POST.get('category')
        price = request.POST.get('price', '').strip()
        product_offer = request.POST.get('product_offer', '0').strip()
        main_image = request.FILES.get('main_image')
        errors = []

        if not product_name:
            errors.append("Product name is required")
        elif len(product_name) < 2:
            errors.append("Product name should be at least 2 characters long")
        elif len(product_name) > 100:
            errors.append("Product name should not exceed 100 characters")
        elif Product.objects.filter(product_name__iexact=product_name, is_deleted=False).exclude(id=product_id).exists():
            errors.append("A product with this name already exists")

        if not category_id:
            errors.append("Please select a category")
        else:
            try:
                category = Category.objects.get(id=category_id, is_deleted=False)
            except Category.DoesNotExist:
                errors.append("Selected category does not exist")

        if not price:
            errors.append("Price is required")
        else:
            try:
                price = int(price)
                if price <= 0:
                    errors.append("Price must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid price")

        try:
            product_offer = float(product_offer)
            if product_offer < 0 or product_offer > 100:
                errors.append("Offer percentage must be between 0 and 100")
        except ValueError:
            errors.append("Please enter a valid offer percentage")

        if main_image:
            if not main_image.content_type.startswith('image/'):
                errors.append("Please upload a valid image file for main image")
            if main_image.size > 10 * 1024 * 1024:
                errors.append("Main image should be less than 10MB")

        if errors:
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            for error in errors:
                messages.error(request, error)
            return render(request, 'edit_product.html', {
                'product': product,
                'categories': categories,
            })

        try:
            product.product_name = product_name
            product.product_description = product_description
            product.category = category
            product.price = price
            product.product_offer = product_offer

            if main_image:
                processed_main_image = resize_and_crop_image(main_image)
                product.main_image.save(
                    f"{product.product_name}_main.jpg",
                    processed_main_image,
                    save=False
                )

            product.save()
            messages.success(request, f"Product '{product_name}' updated successfully")
            return redirect('product_list')
        except Exception as e:
            messages.error(request, "An error occurred while updating the product")
            categories = Category.objects.filter(is_deleted=False).order_by('name')
            return render(request, 'edit_product.html', {
                'product': product,
                'categories': categories,
            })

    categories = Category.objects.filter(is_deleted=False).order_by('name')
    return render(request, 'edit_product.html', {
        'product': product,
        'categories': categories,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_product_status(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        
        action = request.POST.get('action')
        product_name = product.product_name

        if action == 'deactivate' and not product.is_deleted:
            product.is_deleted = True
            product.save()
            messages.success(request, f"Product '{product_name}' has been deactivated successfully")
        elif action == 'activate' and product.is_deleted:
            product.is_deleted = False
            product.save()
            messages.success(request, f"Product '{product_name}' has been activated successfully")
        else:
            messages.error(request, "Invalid action or product status")
        return redirect('product_list')

    return render(request, 'toggle_product_status.html', {
        'product': product,
        
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def product_detail(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    variants = product.variants.prefetch_related('images').all()
    return render(request, 'products_details.html', {
        'product': product,
        'variants': variants
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def product_variants(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    variants = product.variants.prefetch_related('images').all()
    
    return render(request, 'product_variants.html', {
        'product': product,
        'variants': variants
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_variant(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        color_type = request.POST.get('color_type', '').strip()
        color = request.POST.get('color', '').strip()
        color_code = request.POST.get('color_code', '').strip()
        stock_quantity = request.POST.get('stock_quantity', '0').strip()
        images = request.FILES.getlist('images')
        errors = []

        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity < 0:
                errors.append("Stock quantity cannot be negative")
            elif stock_quantity > 9999:
                errors.append("Stock quantity cannot exceed 9999")
        except ValueError:
            errors.append("Please enter a valid stock quantity")

        if color_type == 'predefined':
            if not color:
                errors.append("Please select a predefined color")
            else:
                if ProductVariant.objects.filter(product=product, color=color, color_code__isnull=True).exists():
                    errors.append("A variant with this predefined color already exists for this product")
                color_code = None
        elif color_type == 'custom':
            if not color_code:
                errors.append("Please enter a custom hex color code")
            elif not re.match(r'^#[0-9A-Fa-f]{6}$', color_code):
                errors.append("Please enter a valid hex color code (e.g., #FF0000)")
            else:
                color = 'custom'
                if ProductVariant.objects.filter(product=product, color_code=color_code).exists():
                    errors.append("A variant with this custom color already exists for this product")
        else:
            errors.append("Please select either a predefined color or enter a custom color")

        if len(images) < 3:
            errors.append("Please upload at least 3 images for the variant")
        elif len(images) > 10:
            errors.append("Maximum 10 images allowed per variant")
        else:
            for image in images:
                if not image.content_type.startswith('image/'):
                    errors.append("Please upload only image files")
                    break
                if image.size > 10 * 1024 * 1024:
                    errors.append("Each image should be less than 10MB")
                    break

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'add_variant.html', {
                'product': product,
                'color': color,
                'color_code': color_code,
                'stock_quantity': request.POST.get('stock_quantity', ''),
            })

        try:
            variant = ProductVariant.objects.create(
                product=product,
                color=color,
                color_code=color_code,
                stock_quantity=stock_quantity
            )

            for i, image in enumerate(images):
                try:
                    processed_image = resize_and_crop_image(image)
                    variant_image = VariantImage(
                        variant=variant,
                        is_primary=(i == 0)
                    )
                    variant_image.image.save(
                        f"{product.product_name}_{variant.color}_{i+1}.jpg",
                        processed_image,
                        save=True
                    )
                except Exception as e:
                    messages.warning(request, f"Error processing image {i+1}: {str(e)}")

            if color_type == 'predefined':
                color_display = dict(ProductVariant.COLOR_CHOICES).get(color, color)
                messages.success(request, f"Variant '{color_display}' added successfully with {stock_quantity} items in stock")
            else:
                messages.success(request, f"Custom color variant '{color_code}' added successfully with {stock_quantity} items in stock")
            
            return redirect('product_variants', product_id=product.id)
        except Exception as e:
            messages.error(request, "An error occurred while adding the variant")

    return render(request, 'add_variant.html', {
        'product': product,
        'color_choices': ProductVariant.COLOR_CHOICES
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_variant(request, variant_id):
    variant = get_object_or_404(ProductVariant, id=variant_id)
    
    if request.method == 'POST':
        color_type = request.POST.get('color_type', '').strip()
        color = request.POST.get('color', '').strip()
        color_code = request.POST.get('color_code', '').strip()
        stock_quantity = request.POST.get('stock_quantity', '0').strip()
        images = request.FILES.getlist('images')
        remove_images = request.POST.getlist('remove_images')
        primary_image = request.POST.get('primary_image')
        errors = []

        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity < 0:
                errors.append("Stock quantity cannot be negative")
            elif stock_quantity > 9999:
                errors.append("Stock quantity cannot exceed 9999")
        except ValueError:
            errors.append("Please enter a valid stock quantity")

        if color_type == 'predefined':
            if not color:
                errors.append("Please select a predefined color")
            else:
                if ProductVariant.objects.filter(
                    product=variant.product, 
                    color=color, 
                    color_code__isnull=True
                ).exclude(id=variant.id).exists():
                    errors.append("A variant with this predefined color already exists for this product")
                color_code = None
        elif color_type == 'custom':
            if not color_code:
                errors.append("Please enter a custom hex color code")
            elif not re.match(r'^#[0-9A-Fa-f]{6}$', color_code):
                errors.append("Please enter a valid hex color code (e.g., #FF0000)")
            else:
                color = 'custom'
                if ProductVariant.objects.filter(
                    product=variant.product, 
                    color_code=color_code
                ).exclude(id=variant.id).exists():
                    errors.append("A variant with this custom color already exists for this product")
        else:
            errors.append("Please select either a predefined color or enter a custom color")

        current_images_count = variant.images.count() - len(remove_images)
        total_images_after = current_images_count + len(images)

        if total_images_after < 3:
            errors.append("Variant must have at least 3 images")
        elif total_images_after > 10:
            errors.append("Maximum 10 images allowed per variant")

        if images:
            for image in images:
                if not image.content_type.startswith('image/'):
                    errors.append("Please upload only image files")
                    break
                if image.size > 10 * 1024 * 1024:
                    errors.append("Each image should be less than 10MB")
                    break

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'edit_variant.html', {
                'variant': variant,
                'color_choices': ProductVariant.COLOR_CHOICES
            })

        try:
            variant.color = color
            variant.color_code = color_code
            variant.stock_quantity = stock_quantity
            
            variant.save()

            if remove_images:
                VariantImage.objects.filter(id__in=remove_images, variant=variant).delete()

            for i, image in enumerate(images):
                try:
                    processed_image = resize_and_crop_image(image)
                    variant_image = VariantImage(variant=variant)
                    variant_image.image.save(
                        f"{variant.product.product_name}_{variant.color}_{variant.images.count() + i + 1}.jpg",
                        processed_image,
                        save=True
                    )
                except Exception as e:
                    messages.warning(request, f"Error processing image {i+1}: {str(e)}")

            if primary_image:
                try:
                    VariantImage.objects.filter(variant=variant).update(is_primary=False)
                    VariantImage.objects.filter(id=primary_image, variant=variant).update(is_primary=True)
                except:
                    pass

            if color_type == 'predefined':
                color_display = dict(ProductVariant.COLOR_CHOICES).get(color, color)
                messages.success(request, f"Variant '{color_display}' updated successfully. Stock: {stock_quantity} items")
            else:
                messages.success(request, f"Custom color variant '{color_code}' updated successfully. Stock: {stock_quantity} items")
            
            return redirect('product_variants', product_id=variant.product.id)
        except Exception as e:
            messages.error(request, f"An error occurred while updating the variant =={e}")
            print(e)
    return render(request, 'edit_variant.html', {
        'variant': variant,
        'color_choices': ProductVariant.COLOR_CHOICES
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_variant_status(request, variant_id):
    variant = get_object_or_404(ProductVariant, id=variant_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')

        if variant.color_code:
            variant_color = f"Custom Color ({variant.color_code})"
        else:
            variant_color = variant.get_color_display()
        
        if action == 'deactivate' and variant.is_active:
            variant.is_active = False
            variant.save()
            messages.success(request, f"Variant '{variant_color}' has been deactivated successfully")
        elif action == 'activate' and not variant.is_active:
            variant.is_active = True
            variant.save()
            messages.success(request, f"Variant '{variant_color}' has been activated successfully")
        else:
            messages.error(request, "Invalid action or variant status")
    
    return redirect('product_variants', product_id=variant.product.id)

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def dashboard(request):
    # Get current date and time
    now = timezone.now()
    today = now.date()
    
    # Calculate date ranges
    week_ago = now - timedelta(days=7)
    month_ago = now - timedelta(days=30)
    
    # Basic counts
    total_users = User.objects.count()
    total_products = Product.objects.count()
    total_categories = Category.objects.count()
    total_orders = Order.objects.count()
    
    # Sales statistics
    today_orders = Order.objects.filter(created_at__date=today, status='delivered')
    weekly_orders = Order.objects.filter(created_at__gte=week_ago, status='delivered')
    monthly_orders = Order.objects.filter(created_at__gte=month_ago, status='delivered')
    
    # Revenue calculations
    today_revenue = today_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    weekly_revenue = weekly_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    monthly_revenue = monthly_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_revenue = Order.objects.filter(status='delivered').aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    
    # Order counts
    today_order_count = today_orders.count()
    weekly_order_count = weekly_orders.count()
    monthly_order_count = monthly_orders.count()
    
    # Recent orders
    recent_orders = Order.objects.select_related('user').order_by('-created_at')[:5]
    
    # Top selling products (last 30 days)
    top_products = OrderItem.objects.filter(
        order__created_at__gte=month_ago,
        order__status='delivered'
    ).values(
        'product__product_name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('price')
    ).order_by('-total_quantity')[:5]
    
    # Coupon statistics
    total_coupons = Coupon.objects.count()
    active_coupons = Coupon.objects.filter(is_active=True, valid_until__gte=now).count()
    coupon_usage_today = CouponUsage.objects.filter(used_at__date=today).count()
    total_discount_given = Order.objects.filter(status='delivered').aggregate(
        total=Sum('coupon_discount')
    )['total'] or Decimal('0.00')
    
    # Daily sales data for chart (last 7 days)
    daily_sales = []
    for i in range(7):
        date = (now - timedelta(days=i)).date()
        day_orders = Order.objects.filter(created_at__date=date, status='delivered')
        daily_revenue = day_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        daily_sales.append({
            'date': date.strftime('%Y-%m-%d'),
            'revenue': float(daily_revenue),
            'orders': day_orders.count()
        })
    
    daily_sales.reverse()  # Show oldest to newest
    
    # Order status distribution
    order_status_counts = Order.objects.values('status').annotate(count=Count('id'))
    
    context = {
        'total_users': total_users,
        'total_products': total_products,
        'total_categories': total_categories,
        'total_orders': total_orders,
        'today_revenue': today_revenue,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'total_revenue': total_revenue,
        'today_order_count': today_order_count,
        'weekly_order_count': weekly_order_count,
        'monthly_order_count': monthly_order_count,
        'recent_orders': recent_orders,
        'top_products': top_products,
        'total_coupons': total_coupons,
        'active_coupons': active_coupons,
        'coupon_usage_today': coupon_usage_today,
        'total_discount_given': total_discount_given,
        'daily_sales': daily_sales,
        'order_status_counts': order_status_counts,
    }
    
    return render(request, 'dashboard.html', context)

def admin_profile(request):
    return render(request, 'settings.html')

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def order_management_page(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    payment_filter = request.GET.get('payment', 'all')
    sort_order = request.GET.get('sort', 'desc')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    orders = Order.objects.select_related('user', 'shipping_address').prefetch_related('items__product', 'items__variant')
    
    if query:
        orders = orders.filter(Q(order_number__icontains=query) | Q(user__full_name__icontains=query) | Q(user__email__icontains=query) | Q(items__product__product_name__icontains=query)).distinct()
        
    if status_filter != 'all':
        orders = orders.filter(status=status_filter)
        
    if payment_filter != 'all':
        orders = orders.filter(payment_method=payment_filter)
        
    if date_from:
        try:
            from_date = timezone.datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=from_date)
        except ValueError:
            date_from = ''
            
    if date_to:
        try:
            to_date = timezone.datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=to_date)
        except ValueError:
            date_to = ''
            
    if sort_order == 'asc':
        orders = orders.order_by('created_at')
    else:
        orders = orders.order_by('-created_at')
        
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    delivered_orders = Order.objects.filter(status='delivered').count()
    cancelled_orders = Order.objects.filter(status='cancelled').count()
    total_revenue = Order.objects.filter(status='delivered').aggregate(total=Sum('total_amount'))['total'] or 0
    
    paginator = Paginator(orders, 10)
    page = request.GET.get('page')
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)
        
    context = {
        'orders':orders, 'query': query, 'status_filter': status_filter,
        'payment_filter': payment_filter, 'sort_order': sort_order, 'date_from': date_from,
        'date_to': date_to, 'total_orders': total_orders, 'pending_orders': pending_orders,
        'delivered_orders': delivered_orders, 'cancelled_orders': cancelled_orders, 'total_revenue': total_revenue,
        'order_status_choices': Order.ORDER_STATUS_CHOICES, 'payment_method_choices': Order.PAYMENT_METHOD_CHOICES,
    }
    
    return render(request, 'orders/order_management.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def order_detail_view(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    order_items = order.items.all()
    
    # Calculate item statistics
    active_items_count = order_items.filter(status='active').count()
    cancelled_items_count = order_items.filter(status='cancelled').count()
    returned_items_count = order_items.filter(status='returned').count()
    
    # Get return request if exists (for backward compatibility)
    try:
        return_request = order.return_request
    except:
        return_request = None
    
    context = {
        'order': order,
        'order_items': order_items,
        'return_request': return_request,
        'active_items_count': active_items_count,
        'cancelled_items_count': cancelled_items_count,
        'returned_items_count': returned_items_count,
    }
    
    return render(request, 'orders/orders_details.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def update_order_status(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        
        if new_status in dict(Order.ORDER_STATUS_CHOICES):
            old_status = order.status
            order.status = new_status
            order.save()
            
            if new_status == 'cancelled' and old_status != 'cancelled':
                for item in order.items.all():
                    item.variant.stock_quantity += item.quantity
                    item.variant.save()
                    
            messages.success(request, f"Order {order.order_number} status updated to {order.get_status_display()}")
        else:
            messages.error(request, "Invalid status selected")
            
    return redirect('order_detail_view', order_id=order.id)

@login_required
@user_passes_test(lambda u: u.is_staff)
@require_http_methods(["GET", "POST"])
def verify_return_request(request, return_request_id):
    return_request = get_object_or_404(ReturnRequest, id=return_request_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        admin_notes = request.POST.get('admin_notes', '').strip()
        
        try:
            if action == 'approve':
                if return_request.status != 'pending':
                    messages.error(request, "This return request has already been processed.")
                    return redirect('return_requests_page')

                return_request.status = 'approved'
                return_request.admin_notes = admin_notes
                return_request.processed_at = timezone.now()
                return_request.processed_by = request.user
                return_request.save()

                order = return_request.order
                order.status = 'returned'
                order.save()

                for item in order.items.all():
                    item.variant.stock_quantity += item.quantity
                    item.variant.save()

                wallet, created = Wallet.objects.get_or_create(user=order.user)
                wallet.add_money(
                    order.total_amount,
                    f"Refund for returned order {order.order_number}"
                )
                
                messages.success(
                    request, 
                    f"Return request approved successfully. ₹{order.total_amount} has been added to {order.user.full_name}'s wallet."
                )
                
            elif action == 'reject':
                
                if not admin_notes:
                    messages.error(request, "Admin notes are required when rejecting a return request.")
                    return redirect('return_requests_page')

                if return_request.status != 'pending':
                    messages.error(request, "This return request has already been processed.")
                    return redirect('return_requests_page')

                return_request.status = 'rejected'
                return_request.admin_notes = admin_notes
                return_request.processed_at = timezone.now()
                return_request.processed_by = request.user
                return_request.save()
                
                messages.success(
                    request, 
                    f"Return request for order {return_request.order.order_number} has been rejected."
                )
            
            else:
                messages.error(request, "Invalid action specified.")
                
        except Exception as e:
            messages.error(request, f"An error occurred while processing the return request: {str(e)}")
    
    return redirect('return_requests_page')

@login_required
@user_passes_test(lambda u: u.is_staff)
def return_requests_page(request):
    status_filter = request.GET.get('status', 'all')
    request_type = request.GET.get('type', 'all')  # New filter for request type
    
    # Get both individual item returns and full order returns
    item_returns = ItemReturnRequest.objects.select_related(
        'order_item__order__user', 
        'order_item__product', 
        'order_item__variant',
        'processed_by'
    ).order_by('-requested_at')
    
    full_order_returns = ReturnRequest.objects.select_related(
        'order__user', 
        'processed_by'
    ).order_by('-requested_at')
    
    # Apply status filter
    if status_filter != 'all':
        item_returns = item_returns.filter(status=status_filter)
        full_order_returns = full_order_returns.filter(status=status_filter)
    
    # Apply type filter
    if request_type == 'item':
        full_order_returns = full_order_returns.none()
    elif request_type == 'order':
        item_returns = item_returns.none()
    
    # Combine and sort the querysets
    all_returns = []
    
    # Add item returns
    for item_return in item_returns:
        all_returns.append({
            'type': 'item',
            'id': item_return.id,
            'order_number': item_return.order_item.order.order_number,
            'order_id': item_return.order_item.order.id,
            'customer_name': item_return.order_item.order.user.full_name,
            'customer_email': item_return.order_item.order.user.email,
            'product_name': item_return.order_item.product.product_name,
            'variant_color': item_return.order_item.variant.get_color_display(),
            'quantity': item_return.order_item.quantity,
            'item_total': item_return.order_item.get_total_price(),
            'reason': item_return.reason,
            'status': item_return.status,
            'requested_at': item_return.requested_at,
            'processed_at': item_return.processed_at,
            'processed_by': item_return.processed_by,
            'admin_notes': item_return.admin_notes,
            'object': item_return
        })
    
    # Add full order returns
    for order_return in full_order_returns:
        all_returns.append({
            'type': 'order',
            'id': order_return.id,
            'order_number': order_return.order.order_number,
            'order_id': order_return.order.id,
            'customer_name': order_return.order.user.full_name,
            'customer_email': order_return.order.user.email,
            'product_name': 'Full Order',
            'variant_color': '',
            'quantity': order_return.order.items.count(),
            'item_total': order_return.order.total_amount,
            'reason': order_return.reason,
            'status': order_return.status,
            'requested_at': order_return.requested_at,
            'processed_at': order_return.processed_at,
            'processed_by': order_return.processed_by,
            'admin_notes': order_return.admin_notes,
            'object': order_return
        })
    
    # Sort by requested_at descending
    all_returns.sort(key=lambda x: x['requested_at'], reverse=True)
    
    # Paginate
    paginator = Paginator(all_returns, 10)
    page = request.GET.get('page')
    try:
        return_requests = paginator.page(page)
    except PageNotAnInteger:
        return_requests = paginator.page(1)
    except EmptyPage:
        return_requests = paginator.page(paginator.num_pages)
    
    context = {
        'return_requests': return_requests,
        'status_filter': status_filter,
        'request_type': request_type,
    }
    return render(request, 'orders/return_request.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
@require_http_methods(["GET", "POST"])
def verify_item_return_request(request, item_return_id):
    item_return = get_object_or_404(ItemReturnRequest, id=item_return_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        admin_notes = request.POST.get('admin_notes', '').strip()
        
        try:
            if action == 'approve':
                if item_return.status != 'pending':
                    messages.error(request, "This return request has already been processed.")
                    return redirect('return_requests_page')
                
                # Update return request
                item_return.status = 'approved'
                item_return.admin_notes = admin_notes
                item_return.processed_at = timezone.now()
                item_return.processed_by = request.user
                item_return.save()
                
                # Update order item status
                order_item = item_return.order_item
                order_item.status = 'returned'
                order_item.save()
                
                # Restore stock
                order_item.variant.stock_quantity += order_item.quantity
                order_item.variant.save()
                
                # Add refund to wallet
                wallet, created = Wallet.objects.get_or_create(user=order_item.order.user)
                refund_amount = order_item.get_total_price()
                wallet.add_money(
                    refund_amount,
                    f"Refund for returned item: {order_item.product.product_name} from order {order_item.order.order_number}"
                )
                
                messages.success(
                    request, 
                    f"Item return approved successfully. ₹{refund_amount} has been added to {order_item.order.user.full_name}'s wallet."
                )
                
            elif action == 'reject':
                if not admin_notes:
                    messages.error(request, "Admin notes are required when rejecting a return request.")
                    return redirect('return_requests_page')
                
                if item_return.status != 'pending':
                    messages.error(request, "This return request has already been processed.")
                    return redirect('return_requests_page')
                
                item_return.status = 'rejected'
                item_return.admin_notes = admin_notes
                item_return.processed_at = timezone.now()
                item_return.processed_by = request.user
                item_return.save()
                
                messages.success(
                    request, 
                    f"Item return request for {item_return.order_item.product.product_name} has been rejected."
                )
            else:
                messages.error(request, "Invalid action specified.")
                
        except Exception as e:
            messages.error(request, f"An error occurred while processing the return request: {str(e)}")
    
    return redirect('return_requests_page')

@login_required
@user_passes_test(lambda u: u.is_staff)
def inventory_management(request):
    query = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', 'all')
    stock_filter = request.GET.get('stock', 'all')
    
    variants = ProductVariant.objects.select_related('product__category').filter(product__is_deleted=False, is_active=True)
    
    if query:
        variants = variants.filter(Q(product__product_name__icontains=query) | Q(product__category__name__icontains=query) | Q(color__icontains=query))
        
    if category_filter != 'all':
        try:
            category_filter = int(category_filter)
            variants = variants.filter(product__category_id=category_filter)
        except (ValueError, TypeError):
            category_filter = 'all'
            
    if stock_filter == 'low':
        variants = variants.filter(stock_quantity__lte=5, stock_quantity__gt=0)
    elif stock_filter == 'out':
        variants = variants.filter(stock_quantity=0)
    elif stock_filter == 'available':
        variants = variants.filter(stock_quantity__gt=5)
        
    variants = variants.order_by('stock_quantity', 'product__product_name')
    
    categories = Category.objects.filter(is_deleted=False).order_by('name')
    
    paginator = Paginator(variants, 15)
    page = request.GET.get('page')
    try:
        variants = paginator.page(page)
    except PageNotAnInteger:
        variants = paginator.page(1)
    except EmptyPage:
        variants = paginator.page(paginator.num_pages)
        
    context = {
        'variants': variants, 'categories': categories, 'query': query,
        'category_filter': str(category_filter) if category_filter != 'all' else 'all',
        'stock_filter': stock_filter,
    } 
    return render(request, 'orders/inventory_management.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def update_stock(request, variant_id):
    variant = get_object_or_404(ProductVariant, id=variant_id)
    
    if request.method == 'POST':
        try:
            new_stock = int(request.POST.get('stock_quantity', 0))
            if new_stock >= 0:
                variant.stock_quantity = new_stock
                variant.save()
                messages.success(request, f"Stock updated for {variant.product.product_name} - {variant.get_color_display()}")
            else:
                messages.error(request, "Stock quantity cannot be negative")
        except ValueError:
            messages.error(request, "Invalid stock quantity")
    return redirect('inventory_management')

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def referral_offer_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    
    offers = ReferralOffer.objects.all()
    
    if query:
        offers = offers.filter(
            Q(offer_name__icontains=query) |
            Q(description__icontains=query)
        )
    
    if status_filter == 'active':
        offers = offers.filter(is_active=True)
    elif status_filter == 'inactive':
        offers = offers.filter(is_active=False)
    
    offers = offers.order_by('-created_at')
    
    paginator = Paginator(offers, 10)
    page = request.GET.get('page')
    
    try:
        offers = paginator.page(page)
    except PageNotAnInteger:
        offers = paginator.page(1)
    except EmptyPage:
        offers = paginator.page(paginator.num_pages)
    
    # Statistics
    total_offers = ReferralOffer.objects.count()
    active_offers = ReferralOffer.objects.filter(is_active=True).count()
    total_referrals = ReferralReward.objects.count()
    
    context = {
        'offers': offers,
        'query': query,
        'status_filter': status_filter,
        'total_offers': total_offers,
        'active_offers': active_offers,
        'total_referrals': total_referrals,
    }
    
    return render(request, 'offers/referral_offer_list.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_referral_offer(request):
    if request.method == 'POST':
        offer_name = request.POST.get('offer_name', '').strip()
        description = request.POST.get('description', '').strip()
        reward_type = request.POST.get('reward_type', 'coupon')
        reward_value = request.POST.get('reward_value', '').strip()
        reward_type_detail = request.POST.get('reward_type_detail', 'percentage')
        minimum_order_amount = request.POST.get('minimum_order_amount', '0').strip()
        max_referrals = request.POST.get('max_referrals', '').strip()
        
        errors = []
        
        # Validate offer name
        if not offer_name:
            errors.append("Offer name is required")
        elif len(offer_name) < 3:
            errors.append("Offer name should be at least 3 characters long")
        elif len(offer_name) > 100:
            errors.append("Offer name should not exceed 100 characters")
        
        # Validate reward value
        if not reward_value:
            errors.append("Reward value is required")
        else:
            try:
                reward_value = float(reward_value)
                if reward_value <= 0:
                    errors.append("Reward value must be greater than 0")
                if reward_type_detail == 'percentage' and reward_value > 100:
                    errors.append("Percentage reward cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid reward value")
        
        # Validate minimum order amount
        try:
            minimum_order_amount = float(minimum_order_amount)
            if minimum_order_amount < 0:
                errors.append("Minimum order amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum order amount")
        
        # Validate max referrals
        if max_referrals:
            try:
                max_referrals = int(max_referrals)
                if max_referrals <= 0:
                    errors.append("Maximum referrals must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum referrals number")
        else:
            max_referrals = None
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'offers/add_referral_offer.html', {
                'offer_name': offer_name,
                'description': description,
                'reward_type': reward_type,
                'reward_value': request.POST.get('reward_value', ''),
                'reward_type_detail': reward_type_detail,
                'minimum_order_amount': request.POST.get('minimum_order_amount', ''),
                'max_referrals': request.POST.get('max_referrals', ''),
            })
        
        try:
            ReferralOffer.objects.create(
                offer_name=offer_name,
                description=description,
                reward_type=reward_type,
                reward_value=reward_value,
                reward_type_detail=reward_type_detail,
                minimum_order_amount=minimum_order_amount,
                max_referrals=max_referrals,
                is_active=True
            )
            messages.success(request, f"Referral offer '{offer_name}' created successfully")
            return redirect('referral_offer_list')
        except Exception as e:
            messages.error(request, "An error occurred while creating the offer")
            return render(request, 'offers/add_referral_offer.html', {
                'offer_name': offer_name,
                'description': description,
                'reward_type': reward_type,
                'reward_value': request.POST.get('reward_value', ''),
                'reward_type_detail': reward_type_detail,
                'minimum_order_amount': request.POST.get('minimum_order_amount', ''),
                'max_referrals': request.POST.get('max_referrals', ''),
            })
    
    return render(request, 'offers/add_referral_offer.html')

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_referral_offer(request, offer_id):
    offer = get_object_or_404(ReferralOffer, id=offer_id)
    
    if request.method == 'POST':
        offer_name = request.POST.get('offer_name', '').strip()
        description = request.POST.get('description', '').strip()
        reward_type = request.POST.get('reward_type', 'coupon')
        reward_value = request.POST.get('reward_value', '').strip()
        reward_type_detail = request.POST.get('reward_type_detail', 'percentage')
        minimum_order_amount = request.POST.get('minimum_order_amount', '0').strip()
        max_referrals = request.POST.get('max_referrals', '').strip()
        
        errors = []
        
        # Validate offer name
        if not offer_name:
            errors.append("Offer name is required")
        elif len(offer_name) < 3:
            errors.append("Offer name should be at least 3 characters long")
        elif len(offer_name) > 100:
            errors.append("Offer name should not exceed 100 characters")
        
        # Validate reward value
        if not reward_value:
            errors.append("Reward value is required")
        else:
            try:
                reward_value = float(reward_value)
                if reward_value <= 0:
                    errors.append("Reward value must be greater than 0")
                if reward_type_detail == 'percentage' and reward_value > 100:
                    errors.append("Percentage reward cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid reward value")
        
        # Validate minimum order amount
        try:
            minimum_order_amount = float(minimum_order_amount)
            if minimum_order_amount < 0:
                errors.append("Minimum order amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum order amount")
        
        # Validate max referrals
        if max_referrals:
            try:
                max_referrals = int(max_referrals)
                if max_referrals <= 0:
                    errors.append("Maximum referrals must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum referrals number")
        else:
            max_referrals = None
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'offers/edit_referral_offer.html', {'offer': offer})
        
        try:
            offer.offer_name = offer_name
            offer.description = description
            offer.reward_type = reward_type
            offer.reward_value = reward_value
            offer.reward_type_detail = reward_type_detail
            offer.minimum_order_amount = minimum_order_amount
            offer.max_referrals = max_referrals
            offer.save()
            
            messages.success(request, f"Referral offer '{offer_name}' updated successfully")
            return redirect('referral_offer_list')
        except Exception as e:
            messages.error(request, "An error occurred while updating the offer")
    
    return render(request, 'offers/edit_referral_offer.html', {'offer': offer})

@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_referral_offer_status(request, offer_id):
    offer = get_object_or_404(ReferralOffer, id=offer_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'activate' and not offer.is_active:
            offer.is_active = True
            offer.save()
            messages.success(request, f"Referral offer '{offer.offer_name}' has been activated")
        elif action == 'deactivate' and offer.is_active:
            offer.is_active = False
            offer.save()
            messages.success(request, f"Referral offer '{offer.offer_name}' has been deactivated")
        else:
            messages.error(request, "Invalid action or offer status")
    
    return redirect('referral_offer_list')

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def referral_rewards_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    
    rewards = ReferralReward.objects.select_related(
        'referrer', 'referred_user', 'referral_offer', 'coupon'
    ).all()
    
    if query:
        rewards = rewards.filter(
            Q(referrer__full_name__icontains=query) |
            Q(referrer__email__icontains=query) |
            Q(referred_user__full_name__icontains=query) |
            Q(referred_user__email__icontains=query)
        )
    
    if status_filter == 'claimed':
        rewards = rewards.filter(is_claimed=True)
    elif status_filter == 'unclaimed':
        rewards = rewards.filter(is_claimed=False)
    
    rewards = rewards.order_by('-created_at')
    
    paginator = Paginator(rewards, 15)
    page = request.GET.get('page')
    
    try:
        rewards = paginator.page(page)
    except PageNotAnInteger:
        rewards = paginator.page(1)
    except EmptyPage:
        rewards = paginator.page(paginator.num_pages)
    
    # Statistics
    total_rewards = ReferralReward.objects.count()
    claimed_rewards = ReferralReward.objects.filter(is_claimed=True).count()
    total_reward_value = ReferralReward.objects.aggregate(
        total=Sum('reward_amount')
    )['total'] or 0
    
    context = {
        'rewards': rewards,
        'query': query,
        'status_filter': status_filter,
        'total_rewards': total_rewards,
        'claimed_rewards': claimed_rewards,
        'total_reward_value': total_reward_value,
    }
    
    return render(request, 'offers/referral_rewards_list.html', context)

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def coupon_list(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    discount_type_filter = request.GET.get('discount_type', 'all')
    
    coupons = Coupon.objects.all()
    
    if query:
        coupons = coupons.filter(
            Q(code__icontains=query) |
            Q(description__icontains=query)
        )
    
    if status_filter == 'active':
        coupons = coupons.filter(is_active=True, valid_until__gte=timezone.now())
    elif status_filter == 'inactive':
        coupons = coupons.filter(is_active=False)
    elif status_filter == 'expired':
        coupons = coupons.filter(valid_until__lt=timezone.now())
    
    if discount_type_filter != 'all':
        coupons = coupons.filter(discount_type=discount_type_filter)
    
    coupons = coupons.order_by('-created_at')
    
    # Statistics
    total_coupons = Coupon.objects.count()
    active_coupons = Coupon.objects.filter(is_active=True, valid_until__gte=timezone.now()).count()
    expired_coupons = Coupon.objects.filter(valid_until__lt=timezone.now()).count()
    total_usage = CouponUsage.objects.count()
    
    paginator = Paginator(coupons, 10)
    page = request.GET.get('page')
    
    try:
        coupons = paginator.page(page)
    except PageNotAnInteger:
        coupons = paginator.page(1)
    except EmptyPage:
        coupons = paginator.page(paginator.num_pages)
    
    context = {
        'coupons': coupons,
        'query': query,
        'status_filter': status_filter,
        'discount_type_filter': discount_type_filter,
        'total_coupons': total_coupons,
        'active_coupons': active_coupons,
        'expired_coupons': expired_coupons,
        'total_usage': total_usage,
    }
    
    return render(request, 'coupons/coupon_list.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_coupon(request):
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        discount_type = request.POST.get('discount_type', 'percentage')
        discount_value = request.POST.get('discount_value', '').strip()
        minimum_amount = request.POST.get('minimum_amount', '0').strip()
        maximum_discount = request.POST.get('maximum_discount', '').strip()
        usage_limit = request.POST.get('usage_limit', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        
        errors = []
        
        # Validate coupon code
        if not code:
            errors.append("Coupon code is required")
        elif len(code) < 3:
            errors.append("Coupon code should be at least 3 characters long")
        elif len(code) > 50:
            errors.append("Coupon code should not exceed 50 characters")
        elif not re.match(r'^[A-Z0-9]+$', code):
            errors.append("Coupon code should only contain uppercase letters and numbers")
        elif Coupon.objects.filter(code=code).exists():
            errors.append("A coupon with this code already exists")
        
        # Validate discount value
        if not discount_value:
            errors.append("Discount value is required")
        else:
            try:
                discount_value = float(discount_value)
                if discount_value <= 0:
                    errors.append("Discount value must be greater than 0")
                if discount_type == 'percentage' and discount_value > 100:
                    errors.append("Percentage discount cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid discount value")
        
        # Validate minimum amount
        try:
            minimum_amount = float(minimum_amount)
            if minimum_amount < 0:
                errors.append("Minimum amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum amount")
        
        # Validate maximum discount
        if maximum_discount:
            try:
                maximum_discount = float(maximum_discount)
                if maximum_discount <= 0:
                    errors.append("Maximum discount must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum discount")
        else:
            maximum_discount = None
        
        # Validate usage limit
        if usage_limit:
            try:
                usage_limit = int(usage_limit)
                if usage_limit <= 0:
                    errors.append("Usage limit must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid usage limit")
        else:
            usage_limit = None
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = timezone.make_aware(datetime.strptime(valid_from, '%Y-%m-%dT%H:%M'))
                until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%dT%H:%M'))
                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                
                if until_date <= timezone.now():
                    errors.append("Valid until date must be in the future")
                    
            except ValueError:
                errors.append("Please enter valid dates")
        
        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'coupons/add_coupon.html', {
                'code': request.POST.get('code', ''),
                'description': description,
                'discount_type': discount_type,
                'discount_value': request.POST.get('discount_value', ''),
                'minimum_amount': request.POST.get('minimum_amount', ''),
                'maximum_discount': request.POST.get('maximum_discount', ''),
                'usage_limit': request.POST.get('usage_limit', ''),
                'valid_from': valid_from,
                'valid_until': valid_until,
            })
        
        try:
            Coupon.objects.create(
                code=code,
                description=description,
                discount_type=discount_type,
                discount_value=discount_value,
                minimum_amount=minimum_amount,
                maximum_discount=maximum_discount,
                usage_limit=usage_limit,
                valid_from=from_date,
                valid_until=until_date,
                is_active=True
            )
            messages.success(request, f"Coupon '{code}' created successfully")
            return redirect('coupon_list')
        except Exception as e:
            messages.error(request, "An error occurred while creating the coupon")
            return render(request, 'coupons/add_coupon.html', {
                'code': request.POST.get('code', ''),
                'description': description,
                'discount_type': discount_type,
                'discount_value': request.POST.get('discount_value', ''),
                'minimum_amount': request.POST.get('minimum_amount', ''),
                'maximum_discount': request.POST.get('maximum_discount', ''),
                'usage_limit': request.POST.get('usage_limit', ''),
                'valid_from': valid_from,
                'valid_until': valid_until,
            })
    
    return render(request, 'coupons/add_coupon.html')

@login_required
@user_passes_test(lambda u: u.is_staff)
def edit_coupon(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        discount_type = request.POST.get('discount_type', 'percentage')
        discount_value = request.POST.get('discount_value', '').strip()
        minimum_amount = request.POST.get('minimum_amount', '0').strip()
        maximum_discount = request.POST.get('maximum_discount', '').strip()
        usage_limit = request.POST.get('usage_limit', '').strip()
        valid_from = request.POST.get('valid_from', '').strip()
        valid_until = request.POST.get('valid_until', '').strip()
        is_active = request.POST.get('is_active') == 'on'
        
        errors = []
        
        # Validate coupon code
        if not code:
            errors.append("Coupon code is required")
        elif len(code) < 3:
            errors.append("Coupon code should be at least 3 characters long")
        elif len(code) > 50:
            errors.append("Coupon code should not exceed 50 characters")
        elif not re.match(r'^[A-Z0-9]+$', code):
            errors.append("Coupon code should only contain uppercase letters and numbers")
        elif Coupon.objects.filter(code=code).exclude(id=coupon.id).exists():
            errors.append("A coupon with this code already exists")
        
        # Validate discount value
        if not discount_value:
            errors.append("Discount value is required")
        else:
            try:
                discount_value = float(discount_value)
                if discount_value <= 0:
                    errors.append("Discount value must be greater than 0")
                if discount_type == 'percentage' and discount_value > 100:
                    errors.append("Percentage discount cannot exceed 100%")
            except ValueError:
                errors.append("Please enter a valid discount value")
        
        # Validate minimum amount
        try:
            minimum_amount = float(minimum_amount)
            if minimum_amount < 0:
                errors.append("Minimum amount cannot be negative")
        except ValueError:
            errors.append("Please enter a valid minimum amount")
        
        # Validate maximum discount
        if maximum_discount:
            try:
                maximum_discount = float(maximum_discount)
                if maximum_discount <= 0:
                    errors.append("Maximum discount must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid maximum discount")
        else:
            maximum_discount = None
        
        # Validate usage limit
        if usage_limit:
            try:
                usage_limit = int(usage_limit)
                if usage_limit <= 0:
                    errors.append("Usage limit must be greater than 0")
            except ValueError:
                errors.append("Please enter a valid usage limit")
        else:
            usage_limit = None
        
        # Validate dates
        if not valid_from:
            errors.append("Valid from date is required")
        if not valid_until:
            errors.append("Valid until date is required")
        
        if valid_from and valid_until:
            try:
                from_date = timezone.make_aware(datetime.strptime(valid_from, '%Y-%m-%dT%H:%M'))
                until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%dT%H:%M'))
                
                if from_date >= until_date:
                    errors.append("Valid from date must be before valid until date")
                    
            except ValueError:
                errors.append("Please enter valid dates")
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                coupon.code = code
                coupon.description = description
                coupon.discount_type = discount_type
                coupon.discount_value = discount_value
                coupon.minimum_amount = minimum_amount
                coupon.maximum_discount = maximum_discount
                coupon.usage_limit = usage_limit
                coupon.valid_from = from_date
                coupon.valid_until = until_date
                coupon.is_active = is_active
                coupon.save()
                
                messages.success(request, f"Coupon '{code}' updated successfully")
                return redirect('coupon_list')
            except Exception as e:
                messages.error(request, "An error occurred while updating the coupon")
    
    # Get usage statistics
    usage_count = CouponUsage.objects.filter(coupon=coupon).count()
    
    context = {
        'coupon': coupon,
        'usage_count': usage_count,
    }
    
    return render(request, 'coupons/edit_coupon.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def toggle_coupon_status(request, coupon_id):
    coupon = get_object_or_404(Coupon, id=coupon_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'activate':
            coupon.is_active = True
            coupon.save()
            messages.success(request, f"Coupon '{coupon.code}' has been activated successfully")
        elif action == 'deactivate':
            coupon.is_active = False
            coupon.save()
            messages.success(request, f"Coupon '{coupon.code}' has been deactivated successfully")
        
        return redirect('coupon_list')
    
    # Get usage statistics for confirmation
    usage_count = CouponUsage.objects.filter(coupon=coupon).count()
    
    return render(request, 'coupons/toggle_coupon_status.html', {
        'coupon': coupon,
        'usage_count': usage_count,
    })

@never_cache
@login_required
@user_passes_test(lambda u: u.is_staff)
def sales_report(request):
    # Get filter parameters
    report_type = request.GET.get('report_type', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Calculate date range based on report type
    now = timezone.now()
    
    if report_type == 'daily':
        start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date_obj = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif report_type == 'weekly':
        start_date_obj = now - timedelta(days=7)
        end_date_obj = now
    elif report_type == 'monthly':
        start_date_obj = now - timedelta(days=30)
        end_date_obj = now
    elif report_type == 'yearly':
        start_date_obj = now - timedelta(days=365)
        end_date_obj = now
    elif report_type == 'custom' and start_date and end_date:
        try:
            start_date_obj = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_date_obj = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            messages.error(request, "Invalid date format")
            start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date_obj = now
    else:
        start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date_obj = now
    
    # Get orders within date range
    orders = Order.objects.filter(
        created_at__range=[start_date_obj, end_date_obj],
        status='delivered'  # Only count delivered orders
    ).select_related('user', 'coupon').prefetch_related('items')
    
    # Calculate statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_discount = orders.aggregate(total=Sum('coupon_discount'))['total'] or Decimal('0.00')
    total_shipping = orders.aggregate(total=Sum('shipping_charge'))['total'] or Decimal('0.00')
    
    # Calculate subtotal (revenue + discount - shipping)
    total_subtotal = total_revenue + total_discount - total_shipping
    
    # Get top products
    top_products = OrderItem.objects.filter(
        order__in=orders
    ).values(
        'product__product_name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('price')
    ).order_by('-total_quantity')[:5]
    
    # Get coupon usage statistics
    coupon_usage = orders.exclude(coupon__isnull=True).values(
        'coupon__code'
    ).annotate(
        usage_count=Count('id'),
        total_discount=Sum('coupon_discount')
    ).order_by('-usage_count')[:5]
    
    # Daily sales data for chart (last 7 days)
    daily_sales = []
    for i in range(7):
        date = (now - timedelta(days=i)).date()
        day_orders = Order.objects.filter(created_at__date=date, status='delivered')
        daily_revenue = day_orders.aggregate(total=Sum('total_amount'))['total'] or 0
        daily_sales.append({
            'date': date.strftime('%Y-%m-%d'),
            'revenue': float(daily_revenue),
            'orders': day_orders.count()
        })
    
    daily_sales.reverse()  # Show oldest to newest
    
    context = {
        'report_type': report_type,
        'start_date': start_date_obj.strftime('%Y-%m-%d'),
        'end_date': end_date_obj.strftime('%Y-%m-%d'),
        'start_date_input': start_date,
        'end_date_input': end_date,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_discount': total_discount,
        'total_shipping': total_shipping,
        'total_subtotal': total_subtotal,
        'orders': orders[:20],  # Show latest 20 orders
        'top_products': top_products,
        'coupon_usage': coupon_usage,
        'daily_sales': daily_sales,
    }
    
    return render(request, 'reports/sales_report.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def download_sales_report(request):
    format_type = request.GET.get('format', 'pdf')
    report_type = request.GET.get('report_type', 'daily')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    
    # Calculate date range (same logic as sales_report view)
    now = timezone.now()
    
    if report_type == 'daily':
        start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date_obj = now.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif report_type == 'weekly':
        start_date_obj = now - timedelta(days=7)
        end_date_obj = now
    elif report_type == 'monthly':
        start_date_obj = now - timedelta(days=30)
        end_date_obj = now
    elif report_type == 'yearly':
        start_date_obj = now - timedelta(days=365)
        end_date_obj = now
    elif report_type == 'custom' and start_date and end_date:
        try:
            start_date_obj = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            end_date_obj = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date_obj = now
    else:
        start_date_obj = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date_obj = now
    
    # Get data
    orders = Order.objects.filter(
        created_at__range=[start_date_obj, end_date_obj],
        status='delivered'
    ).select_related('user', 'coupon').prefetch_related('items')
    
    # Calculate statistics
    total_orders = orders.count()
    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    total_discount = orders.aggregate(total=Sum('coupon_discount'))['total'] or Decimal('0.00')
    total_shipping = orders.aggregate(total=Sum('shipping_charge'))['total'] or Decimal('0.00')
    total_subtotal = total_revenue + total_discount - total_shipping
    
    if format_type == 'excel':
        return generate_excel_report(orders, {
            'report_type': report_type,
            'start_date': start_date_obj,
            'end_date': end_date_obj,
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'total_discount': total_discount,
            'total_shipping': total_shipping,
            'total_subtotal': total_subtotal,
        })
    else:
        return generate_pdf_report(orders, {
            'report_type': report_type,
            'start_date': start_date_obj,
            'end_date': end_date_obj,
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'total_discount': total_discount,
            'total_shipping': total_shipping,
            'total_subtotal': total_subtotal,
        })

def generate_pdf_report(orders, stats):
    template_path = 'reports/sales_report_pdf.html'
    context = {
        'orders': orders,
        'stats': stats,
        'generated_at': timezone.now(),
    }
    
    template = get_template(template_path)
    html = template.render(context)
    
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        filename = f'sales_report_{stats["report_type"]}_{stats["start_date"].strftime("%Y%m%d")}_to_{stats["end_date"].strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    return HttpResponse("Error generating PDF", status=500)

def generate_excel_report(orders, stats):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="16213e", end_color="16213e", fill_type="solid")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f'Sales Report - {stats["report_type"].title()}'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Date range
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Period: {stats["start_date"].strftime("%Y-%m-%d")} to {stats["end_date"].strftime("%Y-%m-%d")}'
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Summary statistics
    ws['A4'] = 'Summary Statistics'
    ws['A4'].font = Font(bold=True, size=14)
    
    ws['A5'] = 'Total Orders:'
    ws['B5'] = stats['total_orders']
    ws['A6'] = 'Total Revenue:'
    ws['B6'] = float(stats['total_revenue'])
    ws['A7'] = 'Total Discount:'
    ws['B7'] = float(stats['total_discount'])
    ws['A8'] = 'Total Shipping:'
    ws['B8'] = float(stats['total_shipping'])
    ws['A9'] = 'Subtotal:'
    ws['B9'] = float(stats['total_subtotal'])
    
    # Orders table header
    ws['A11'] = 'Order Details'
    ws['A11'].font = Font(bold=True, size=14)
    
    headers = ['Order Number', 'Customer', 'Date', 'Subtotal', 'Discount', 'Shipping', 'Total', 'Coupon']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Orders data
    for row, order in enumerate(orders, 14):
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.user.full_name)
        ws.cell(row=row, column=3, value=order.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=4, value=float(order.subtotal))
        ws.cell(row=row, column=5, value=float(order.coupon_discount))
        ws.cell(row=row, column=6, value=float(order.shipping_charge))
        ws.cell(row=row, column=7, value=float(order.total_amount))
        ws.cell(row=row, column=8, value=order.coupon.code if order.coupon else 'None')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'sales_report_{stats["report_type"]}_{stats["start_date"].strftime("%Y%m%d")}_to_{stats["end_date"].strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response